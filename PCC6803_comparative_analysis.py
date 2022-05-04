#Module List
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, colors, Border, Side, PatternFill, GradientFill
import pandas as pd
import numpy as np
import subprocess
import urllib.request
global data_url
import time
import re

# THIS CODE USES CENTRAL METABOLISM TABLE (ST1)

#Get index positions of value in dataframe
def getIndexes_row(dfObj, value):
    ''' Get index positions of value in dataframe i.e. dfObj.'''
    listOfPos = list()

    # Get bool dataframe with True at positions where the given value exists
    result = dfObj.isin([value])

    # Get list of columns that contains the value
    seriesObj = result.any()
    columnNames = list(seriesObj[seriesObj == True].index)

    # Iterate over list of columns and fetch the rows indexes where value exists
    for col in columnNames:
        rows = list(result[col][result[col] == True].index)
        for row in rows:
            listOfPos.append((row, col))

    # Return a list of tuples indicating the positions of value in the dataframe
    b = [i[0] for i in listOfPos]

    return b


# 1. Load in the data
# Opens up an edited version of Supplementary Table 1 from Mills et al. (2020)
# 'Current knowledge and recent advances in understanding metabolism of the model
# cyanobacterium Synechocystis sp. PCC 6803'
# Change the input here to any of the other files if of more use.
filename = 'input_files\ST1-PCC-6803-central-metabolism.xlsx'
wb = openpyxl.load_workbook(filename)
ws = wb.active

# Opens up an edited version of Supplementary Table 3 from Baers et al. (2019)
# 'Proteome Mapping of a Cyanobacterium Reveals Distinct Compartment Organization
# and Cell-Dispersed Metabolism'
filename2 = 'input_files\ST3_Baers-SuppTable3_all-PCC-6803-proteins.xlsx'
wb2 = openpyxl.load_workbook(filename2)
ws2 = wb2.active

# 2. Converts excel files into a pd dataframes and np arrays

# a) First we will convert PCC_6803_central_metabolism_input.xlsx
# converts PCC_6803_central_metabolism_input.xlsx into a pandas dataframe for ease of data manipulation.
df_cent_met = pd.read_excel(filename)

# replaces all NaN with @ symbol as arrays do not like 'blank' Indexes.
df_cent_met = df_cent_met.replace(np.nan, '@', regex=True)

# turns our df_cent_met into np.array
arr_cent_met = df_cent_met.to_numpy()
#creates a dataframe with only our first column called KEGG ID
arr_cent_met_col_1 = df_cent_met['KEGG ID '].to_numpy()

# b) now we will do the same to PCC_6803_Baers_SuppTable3_all_proteins.xlsx
# turn into a dataframe and fill all blank spaces with '@'
df_input = pd.read_excel(filename2)
df_input['Blank'] = '@'
df_input = df_input.replace(np.nan, '@', regex=True)

# We only require some information from this file, so we create a reduced df for input
df_input_extra = df_input.filter(['Accession', 'Functional Sub-Category', 'Functional Category', 'Svm Score', 'Blank'],
                                 axis=1)
# Then we convert to an array
arr_input_extra = df_input_extra.to_numpy()


# 3. Now we aim to concatenate the two arrays in order to combine the information ready to be blasted later on, but
# first we manipulate the data to match it's correspondant in the extra information data set.
concat_list = []
col_head = ['Accession', 'Uniprot Accession', 'Gene Product', 'Gene Name', 'Gene names in Literature',
            'Functional Sub-Category', 'Functional Category', 'Localisation', 'Svm Score', 'MW (kDa)', 'No. of TMHs',
            '@']

# Function which can turn an array into a list
def list_from_array(Array):
    array_list = Array.tolist()
    return array_list

# Function which can turn an array into an excel file
def array_to_excel(array, worksheet, row_start, col_start):
    for x in range(0, len(array)):

        data_ls = list(array[x])
        for y in range(0, len(data_ls)):
            worksheet.cell(row=(row_start + x), column=(col_start + y)).value = data_ls[y]

for value in arr_cent_met_col_1:
    # takes each index in arr_cent_met_col_1 and adds to list x
    x = value
    for value2 in arr_input_extra[:, 0]:
    # unpacks all the KEGG IDs from arr_input_extra in column 0 and attempts to match them up with
        y = value2
        if x == y:
            #get the index position from df_cent_met and x
            Cent_Met_Rows = getIndexes_row(df_cent_met, x)
            #get the values in the corresponding index positions in the array
            Cent_Met_Values = arr_cent_met[Cent_Met_Rows, :]
            # converts all these to a value within a list
            First_values = list_from_array(Cent_Met_Values)

            concat_list.append(First_values[0])

            '----------------------------------------------------------------'
            # does the same as above but for the extra information from Baers et al. Table
            New_rows = getIndexes_row(df_input_extra, y)
            New_rows_values = arr_input_extra[New_rows, 1:5]
            Second_values = list_from_array(New_rows_values)

            concat_list.append(Second_values[0])

# takes the concatenated list created above and uses the extend and pop function to see each element
# as it's own within a list
array_list = []
while concat_list:
    array_list.extend(concat_list.pop(0))

# 4. We will now create our concatenated input array
length = int(len(array_list))
result_array = np.asarray(array_list)
result_array = result_array.reshape(int(length / 12), 12)

# and convert to a pd dataframe and add header titles to be able to convert to .txt file
df_concatenated_input = pd.DataFrame(result_array, columns=['Accession', 'Uniprot Accession', 'Gene Product',
                                                            'Gene Name', 'Gene names in Literature', 'Localisation',
                                                            'MW (kDa)', 'No. of TMHs', 'Functional Sub-Category',
                                                            'Functional Category', 'Svm Score', '@'])
#swaps the column order for correct output
df_concatenated_input = df_concatenated_input[['Accession', 'Uniprot Accession', 'Gene Product', 'Gene Name',
                                               'Gene names in Literature', 'Functional Sub-Category',
                                               'Functional Category', 'Localisation', 'Svm Score', 'MW (kDa)',
                                               'No. of TMHs', '@']]

#creates the .txt file and prints to input_files as this is still an input file
df_concatenated_input.to_csv('input_files\concatenated_input.txt', index=False, sep='\t')




# 5. Upload files for BLAST-ing
# exact same as the input file created above
input = open('input_files\concatenated_input.txt', 'r')

# unpacks the uniprot ID from the input file
uniprot = []
for line in input:
    if line.startswith('Accession'):
        pass
    else:
        line = line.split('\t')
        uniprot.append(line[1])

input_df = pd.read_csv('input_files\concatenated_input.txt', sep= '\t')
# PCC 11901 fasta file of protein
accession = 'input_files\Syn11901.faa'

Syn11901_db = str(subprocess.getoutput('makeblastdb -in input_files\Syn11901.faa -dbtype prot'))

# function which takes Uniprot ids and can search against the Uniprot database
from bioservices import UniProt
u = UniProt(verbose=False)


row_for_blasting = []
output = []
counter = 0 

# 6. Create an empty dataframe for the BLAST output/results to go in
df = pd.DataFrame()
# with column headers called:
headers = ['BLAST Output', 'PCC 11901 Accession', 'Percentage Identity', 'Alignment Length',
           'Mismatches', 'Gap Open', 'Synechocystis sp. PCC 6803 Start', 'Synechocystis sp. PCC 6803 End',
           'PCC 11901 Start', 'PCC 11901 End', 'E-value', 'Bit Score']

# we will also make a blank array with the same column headers
headers = np.asarray(headers)

# now we attach the sequence to the appropriate Uniprot id
no_blast = []
for id in uniprot:
    if id == 'Uniprot Accession':
        pass
    else:
        if row_for_blasting == []: # if line in input is blank, also ignore
            pass
        else:
            output.append("\t".join(row_for_blasting)) # if there is data there, append the row of data to the temporary row
        row_for_blasting = [] # reinitialise for the loop
        id = str(id) # turn the uniprot id into a string
        fasta = u.search(id, frmt='fasta') # use the bioservice previously imported to search the uniprot id and
        # retrieve the fasta data accesses the fasta file from the website for each protein.
        sequence = fasta.split("\n")[1:]
        sequence = "".join(sequence)
        blasting = open("blasting.fa", "w") # creates a temporary fast file required
        print(fasta.split("\n")[0], sequence, sep="\n" , file= blasting) # appends the sequence or blasting to a
        # temporary fasta file that we can call for the bBLAST subprocess and then cycle through and replace each time.
        blasting.close()
        # the all important BLAST line using the computer's console to compute
        blasted = str(subprocess.getoutput('blastp -outfmt 7 -query blasting.fa -db input_files\Syn11901.faa -evalue 1'))
        blasted = blasted.split('hits found') # formats output
        list_each = []
        for each in blasted:
            # This loop is used to format the blast output
            each = str(each)
            if each.startswith('#') or 'BLASTP' in each or each == '\'':
                pass
            else:
                each = each.replace('\\r', '\\t')
                each = each.replace('\n', '\t')
                # a and b are used for reformatting to fit into the dataframe
                a = [each]
                b = a.pop(-1)
                each = each.split('\t')
                each = each[:-1]
                del each[0]
                print(len(each) / 12) # confirms that each output can be divided by 12 to fit into the array
                # after reformatting
                head = np.array(input_df.iloc[counter, :])
                each_array = np.array(each) # turns each into an array
                results = np.concatenate((head, headers), axis=0) # concatenates the to arrays (head and headers)
                results = np.concatenate((results, each_array), axis=0)  # concatenates the to array just created
                # (results) with the array of blast outputs
                length = int(len(results) / 12) # breaks the BLAST output so that each hit is on a new line
                results = results.reshape(length, 12)
                print('potato', np.shape(results)) #check that it is iterating through

                df = df.append(pd.DataFrame(results), ignore_index=True)
                counter = counter + 1 # counter helps to keep track
                break

df = df.replace('@', ' ', regex=True)

wb = Workbook()
global results_file
timestr = time.strftime('%Y_%m_%d-%H_%M')
results_file = 'output_files\{}_cent-met.xlsx'.format(timestr) # You can rename with an appropriate output name here

df.to_excel(results_file, index=False)

# 7. Formats the output

wb = openpyxl.load_workbook(results_file)
ws = wb.active

#Creates sheet for the Key
ws_key = wb.create_sheet('Key')
ws_key.title = 'Key'

df = pd.read_excel(results_file)

def getIndexes(dfObj, value):
    ''' Get index positions of value in dataframe i.e. dfObj.'''
    listOfPos = list()
    # Get bool dataframe with True at positions where the given value exists
    result = dfObj.isin([value])
    # Get list of columns that contains the value
    seriesObj = result.any()
    columnNames = list(seriesObj[seriesObj == True].index)
    # Iterate over list of columns and fetch the rows indexes where value exists
    for col in columnNames:
        rows = list(result[col][result[col] == True].index)
        for row in rows:
            listOfPos.append((row, col))
    # Return a list of tuples indicating the positions of value in the dataframe
    return listOfPos

#calls function above
a = getIndexes(df, 'BLAST Output')

#takes the first index postion for rows
b = [i[0] for i in a]
#adds one to the index position for excel to read. Gets KEGG id for blasted protein
c = [c+1 for c in b]
#adds one to the index position for excel to read. Gets BLAST Output line in
d = [d+1 for d in c]

thick_border = Border(bottom=Side(style="thick", color='00000000')) #create border style

# Create a NamedStyle (if not already defined)
if 'blast_output_header' not in wb.named_styles:
    blast_output_header = NamedStyle(name='blast_output_header')
    blast_output_header.font = Font(color='00000099', italic=True) #grey and italicised
    blast_output_header.Border = thick_border
    #blast_output_header.font = Font(color='00FF0000', italic=True) #red
    wb.add_named_style(blast_output_header)

if 'blast_input_header' not in wb.named_styles:
    blast_input_header = NamedStyle(name='blast_input_header')
    blast_input_header.font = Font(color='00000000', bold=True) #create_font_style
    #blast_input_header.border = Border(bottom= thick) #add border to cells
    wb.add_named_style(blast_input_header)

if 'main_headers' not in wb.named_styles:
    main_headers = NamedStyle(name='main_headers')
    main_headers.fill = PatternFill('solid', fgColor='00CCE5FF') #cell colour
    main_headers.font = Font(color='00000000', bold=True, size = 12)  # black and bold
    wb.add_named_style(main_headers)

if 'Amino acid biosynthesis' not in wb.named_styles:
    Amino_acid_biosynthesis = NamedStyle(name='Amino acid biosynthesis')
    Amino_acid_biosynthesis.fill = PatternFill('solid', fgColor='00FF6666') #cell colour
    Amino_acid_biosynthesis.font = Font(color='00000000', bold=True) #black and bold
    wb.add_named_style(Amino_acid_biosynthesis)
if 'Biosynthesis of cofactors, prosthetic groups, and carriers' not in wb.named_styles:
    Biosynthesis_of_cofactors_prosthetic_groups_and_carriers = NamedStyle(name='Biosynthesis of cofactors, prosthetic groups, and carriers')
    Biosynthesis_of_cofactors_prosthetic_groups_and_carriers.fill = PatternFill('solid', fgColor='00FFB266') #cell colour
    Biosynthesis_of_cofactors_prosthetic_groups_and_carriers.font = Font(color='00000000', bold=True)  # black and bold
    wb.add_named_style(Biosynthesis_of_cofactors_prosthetic_groups_and_carriers)
if 'Cell Envelope' not in wb.named_styles:
    Cell_Envelope = NamedStyle(name='Cell Envelope')
    Cell_Envelope.fill = PatternFill('solid', fgColor='00FFFF66') #cell colour
    Cell_Envelope.font = Font(color='00000000', bold=True)  # black and bold
    wb.add_named_style(Cell_Envelope)
if 'Cellular Processes' not in wb.named_styles:
    Cellular_Processes = NamedStyle(name='Cellular Processes')
    Cellular_Processes.fill = PatternFill('solid', fgColor='00B2FF66') #cell colour
    Cellular_Processes.font = Font(color='00000000', bold=True) #black and bold
    wb.add_named_style(Cellular_Processes)
if 'Central Intermediary Metabolism' not in wb.named_styles:
    Central_Intermediary_Metabolism = NamedStyle(name='Central Intermediary Metabolism')
    Central_Intermediary_Metabolism.fill = PatternFill('solid', fgColor='0066FF66') #cell colour
    Central_Intermediary_Metabolism.font = Font(color='00000000', bold=True) #black and bold
    wb.add_named_style(Central_Intermediary_Metabolism)
if 'DNA replication, restriction, modification, recombination, and repair' not in wb.named_styles:
    DNA_replication_restriction_modification_recombination_and_repair = NamedStyle(name='DNA replication, restriction, modification, recombination, and repair')
    DNA_replication_restriction_modification_recombination_and_repair.fill = PatternFill('solid', fgColor='0066FFB2') #cell colour
    DNA_replication_restriction_modification_recombination_and_repair.font = Font(color='00000000', bold=True)  # black and bold
    wb.add_named_style(DNA_replication_restriction_modification_recombination_and_repair)
if 'Energy Metabolism' not in wb.named_styles:
    Energy_Metabolism = NamedStyle(name='Energy Metabolism')
    Energy_Metabolism.fill = PatternFill('solid', fgColor='0066FFFF') #cell colour
    Energy_Metabolism.font = Font(color='00000000', bold=True) #black and bold
    wb.add_named_style(Energy_Metabolism)
if 'Fatty acid, phospholipid and sterol metabolism' not in wb.named_styles:
    Fatty_acid_phospholipid_and_sterol_metabolism = NamedStyle(name='Fatty acid, phospholipid and sterol metabolism')
    Fatty_acid_phospholipid_and_sterol_metabolism.fill = PatternFill('solid', fgColor='0066B2FF') #cell colour
    Fatty_acid_phospholipid_and_sterol_metabolism.font = Font(color='00000000', bold=True) #black and bold
    wb.add_named_style(Fatty_acid_phospholipid_and_sterol_metabolism)
if 'Other Categories' not in wb.named_styles:
    Other_Categories = NamedStyle(name='Other Categories')
    Other_Categories.fill = PatternFill('solid', fgColor='006666FF') #cell colour
    Other_Categories.font = Font(color='00000000', bold=True) #black and bold
    wb.add_named_style(Other_Categories)
if 'Photosynthesis & Respiration' not in wb.named_styles:
    Photosynthesis_and_Respiration = NamedStyle(name='Photosynthesis & Respiration')
    Photosynthesis_and_Respiration.fill = PatternFill('solid', fgColor='00B266FF') #cell colour
    Photosynthesis_and_Respiration.font = Font(color='00000000', bold=True) #black and bold
    wb.add_named_style(Photosynthesis_and_Respiration)
if 'Purines, pyrimidines, nucleosides, and nucleotides' not in wb.named_styles:
    Purines_pyrimidines_nucleosides_and_nucleotides = NamedStyle(name='Purines, pyrimidines, nucleosides, and nucleotides')
    Purines_pyrimidines_nucleosides_and_nucleotides.fill = PatternFill('solid', fgColor='00B266FF') #cell colour
    Purines_pyrimidines_nucleosides_and_nucleotides.font = Font(color='00000000', bold=True)  # black and bold
    wb.add_named_style(Purines_pyrimidines_nucleosides_and_nucleotides)
if 'Random proteins (Hypothetical)' not in wb.named_styles:
    Random_proteins_Hypothetical = NamedStyle(name='Random proteins (Hypothetical)')
    Random_proteins_Hypothetical.fill = PatternFill('solid', fgColor='00FF66FF') #cell colour
    Random_proteins_Hypothetical.font = Font(color='00000000', bold=True) #black and bold
    wb.add_named_style(Random_proteins_Hypothetical)
if 'Regulatory Functions' not in wb.named_styles:
    Regulatory_Functions = NamedStyle(name='Regulatory Functions')
    Regulatory_Functions.fill = PatternFill('solid', fgColor='00FF9999') #cell colour
    Regulatory_Functions.font = Font(color='00000000', bold=True) #black and bold
    wb.add_named_style(Regulatory_Functions)
if 'Transcription / Translation' not in wb.named_styles:
    Transcription_Translation = NamedStyle(name='Transcription / Translation')
    Transcription_Translation.fill = PatternFill('solid', fgColor='00E5FFCC') #cell colour
    Transcription_Translation.font = Font(color='00000000', bold=True) #black and bold
    wb.add_named_style(Transcription_Translation)
if 'Transport & Binding Proteins' not in wb.named_styles:
    Transport_and_Binding_Proteins = NamedStyle(name='Transport & Binding Proteins')
    Transport_and_Binding_Proteins.fill = PatternFill('solid', fgColor='00CCFFFF') #cell colour
    Transport_and_Binding_Proteins.font = Font(color='00000000', bold=True) #black and bold
    wb.add_named_style(Transport_and_Binding_Proteins)
if 'Unknown Proteins' not in wb.named_styles:
    Unknown_Proteins = NamedStyle(name='Unknown Proteins')
    Unknown_Proteins.fill = PatternFill('solid', fgColor='00E5CCFF') #cell colour
    Unknown_Proteins.font = Font(color='00000000', bold=True) #black and bold
    wb.add_named_style(Unknown_Proteins)


#change font styles depending on row
for x in d:
    for row in ws[x]:
        row.style = 'blast_output_header'

for y in c:
    for row in ws[y]:
        row.style = 'blast_input_header'


# change cell colours depending on functional categories

def get_cat(dataframe, search):

    a = getIndexes(dataframe, search)

    # takes the first index position for rows
    b = [i[0] for i in a]

   # adds one to the index position for excel to read. Gets KEGG id for blasted protein
    c = [c + 1 for c in b]

    # adds one to the index position for excel to read. Gets BLAST Output line in
    d = [d + 1 for d in c]

    return d

Amino_acid_biosynthesis = get_cat(df, 'Amino acid biosynthesis')
for x in Amino_acid_biosynthesis:
    for row in ws[x]:
        row.style = 'Amino acid biosynthesis'
Biosynthesis_of_cofactors_prosthetic_groups_and_carriers = get_cat(df, 'Biosynthesis of cofactors, prosthetic groups, and carriers')
for x in Biosynthesis_of_cofactors_prosthetic_groups_and_carriers:
    for row in ws[x]:
        row.style = 'Biosynthesis of cofactors, prosthetic groups, and carriers'
Cell_Envelope = get_cat(df, 'Cell Envelope')
for x in Cell_Envelope:
    for row in ws[x]:
        row.style = 'Cell Envelope'
Cellular_Processes = get_cat(df, 'Cellular Processes')
for x in Cellular_Processes:
    for row in ws[x]:
        row.style = 'Cellular Processes'
Central_Intermediary_Metabolism = get_cat(df, 'Central Intermediary Metabolism')
for x in Central_Intermediary_Metabolism:
    for row in ws[x]:
        row.style = 'Central Intermediary Metabolism'
DNA_replication_restriction_modification_recombination_and_repair = get_cat(df, 'DNA replication, restriction, modification, recombination, and repair')
for x in DNA_replication_restriction_modification_recombination_and_repair:
    for row in ws[x]:
        row.style = 'DNA replication, restriction, modification, recombination, and repair'
Energy_Metabolism = get_cat(df, 'Energy Metabolism')
for x in Energy_Metabolism:
    for row in ws[x]:
        row.style = 'Energy Metabolism'
Fatty_acid_phospholipid_and_sterol_metabolism = get_cat(df, 'Fatty acid, phospholipid and sterol metabolism')
for x in Fatty_acid_phospholipid_and_sterol_metabolism:
    for row in ws[x]:
        row.style = 'Fatty acid, phospholipid and sterol metabolism'
Other_Categories = get_cat(df, 'Other Categories')
for x in Other_Categories:
    for row in ws[x]:
        row.style = 'Other Categories'
Photosynthesis_and_Respiration = get_cat(df, 'Photosynthesis & Respiration')
for x in Photosynthesis_and_Respiration:
    for row in ws[x]:
        row.style = 'Photosynthesis & Respiration'
Purines_pyrimidines_nucleosides_and_nucleotides = get_cat(df, 'Purines, pyrimidines, nucleosides, and nucleotides')
for x in Purines_pyrimidines_nucleosides_and_nucleotides:
    for row in ws[x]:
        row.style = 'Purines, pyrimidines, nucleosides, and nucleotides'
Random_proteins_Hypothetical = get_cat(df, 'Random proteins (Hypothetical)')
for x in Random_proteins_Hypothetical:
    for row in ws[x]:
        row.style = 'Random proteins (Hypothetical)'
Regulatory_Functions = get_cat(df, 'Regulatory Functions')
for x in Regulatory_Functions:
    for row in ws[x]:
        row.style = 'Regulatory Functions'
Transcription_Translation = get_cat(df, 'Transcription / Translation')
for x in Transcription_Translation:
    for row in ws[x]:
        row.style = 'Transcription / Translation'
Transport_and_Binding_Proteins = get_cat(df, 'Transport & Binding Proteins')
for x in Transport_and_Binding_Proteins:
    for row in ws[x]:
        row.style = 'Transport & Binding Proteins'
Unknown_Proteins = get_cat(df, 'Unknown Proteins')
for x in Unknown_Proteins:
    for row in ws[x]:
        row.style = 'Unknown Proteins'

#setting the column width
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 17
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 25
ws.column_dimensions['G'].width = 27
ws.column_dimensions['H'].width = 27
ws.column_dimensions['I'].width = 27
ws.column_dimensions['J'].width = 27
ws.column_dimensions['K'].width = 11
ws.column_dimensions['L'].width = 11

#adds and edits column headers to excel sheet
col_head = ['Accession', 'Uniprot Accession', 'Gene Product', 'Gene Name', 'Gene names in Literature',
            'Functional Sub-Category', 'Functional Category', 'Localisation', 'Svm Score', 'MW (kDa)',
            'No. of TMHs', '#']
for x in range(1, len(col_head)):
    (ws.cell(row=1, column=x).value) = col_head[x-1]

#styles and freezes the header columns
for row in ws[1]:
    row.style = 'main_headers'
ws.freeze_panes = 'A2'

#Creates key in another worksheet and colour coordinates.
ws_key["A1"] = 'Amino acid biosynthesis'
ws_key.column_dimensions['A'].width = 57
ws_key["B1"].style = 'Amino acid biosynthesis'
ws_key["A2"] = 'Biosynthesis of cofactors, prosthetic groups, and carriers'
ws_key["B2"].style = 'Biosynthesis of cofactors, prosthetic groups, and carriers'
ws_key["A3"] = 'Cell Envelope'
ws_key["B3"].style = 'Cell Envelope'
ws_key["A4"] = 'Cellular Processes'
ws_key["B4"].style = 'Cellular Processes'
ws_key["A5"] = 'Central Intermediary Metabolism'
ws_key["B5"].style = 'Central Intermediary Metabolism'
ws_key["A6"] = 'DNA replication, restriction, modification, recombination, and repair'
ws_key["B6"].style = 'DNA replication, restriction, modification, recombination, and repair'
ws_key["A7"] = 'Energy Metabolism'
ws_key["B7"].style = 'Energy Metabolism'
ws_key["A8"] = 'Fatty acid, phospholipid and sterol metabolism'
ws_key["B8"].style = 'Fatty acid, phospholipid and sterol metabolism'
ws_key["A9"] = 'Other Categories'
ws_key["B9"].style = 'Other Categories'
ws_key["A10"] = 'Photosynthesis & Respiration'
ws_key["B10"].style = 'Photosynthesis & Respiration'
ws_key["A11"] = 'Purines, pyrimidines, nucleosides, and nucleotides'
ws_key["B11"].style = 'Purines, pyrimidines, nucleosides, and nucleotides'
ws_key["A12"] = 'Random proteins (Hypothetical)'
ws_key["B12"].style = 'Random proteins (Hypothetical)'
ws_key["A13"] = 'Regulatory Functions'
ws_key["B13"].style = 'Regulatory Functions'
ws_key["A14"] = 'Transcription / Translation'
ws_key["B14"].style = 'Transcription / Translation'
ws_key["A15"] = 'Transport & Binding Proteins'
ws_key["B15"].style = 'Transport & Binding Proteins'
ws_key["A16"] = 'Unknown Proteins'
ws_key["B16"].style = 'Unknown Proteins'


wb.save(filename='output_files\{}_cent-met_edited.xlsx.'.format(timestr))#again, one can change the name of the output to suit their format
